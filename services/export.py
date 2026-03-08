"""
Excel导出服务
"""
import os
from datetime import datetime
from typing import Optional, List
from sqlalchemy.orm import Session

from core.database import get_logger
from api.models import Supplier
from services.supplier import SupplierService


logger = get_logger(__name__)


class ExportService:
    """Excel导出服务"""

    @staticmethod
    def export_to_excel(
        db: Session,
        date_filter: Optional[datetime] = None,
        keyword: Optional[str] = None,
        company_type: Optional[str] = None,
        city: Optional[str] = None,
        tags: Optional[List[str]] = None
    ) -> bytes:
        """
        将供应商数据导出为Excel文件

        Args:
            db: 数据库会话
            date_filter: 日期筛选
            keyword: 关键词搜索
            company_type: 公司类型筛选
            city: 城市筛选
            tags: 标签筛选

        Returns:
            bytes: Excel文件的字节数据
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        logger.info("开始导出Excel文件...")

        # 获取数据
        try:
            suppliers, total = SupplierService.get_all(
                db=db,
                skip=0,
                limit=100000,  # 导出所有数据
                date_filter=date_filter,
                keyword=keyword,
                company_type=company_type,
                city=city,
                tags=tags
            )
            logger.info(f"找到 {total} 条记录")

        except Exception as e:
            logger.error(f"获取数据失败: {e}")
            raise

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "供应商数据"

        # 定义表头
        headers = [
            "ID", "公司名称", "URL", "公司类型", "城市",
            "国家", "邮箱", "电话", "标签", "创建时间", "更新时间"
        ]

        # 设置表头样式
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置行高
        ws.row_dimensions[1].height = 25

        # 设置列宽
        column_widths = {
            "A": 8,    # ID
            "B": 30,   # 公司名称
            "C": 40,   # URL
            "D": 15,   # 公司类型
            "E": 15,   # 城市
            "F": 15,   # 国家
            "G": 25,   # 邮箱
            "H": 20,   # 电话
            "I": 20,   # 标签
            "J": 20,   # 创建时间
            "K": 20    # 更新时间
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 写入数据
        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        for row_num, supplier in enumerate(suppliers, 2):
            # 处理标签（将列表转换为字符串）
            tags_str = ", ".join(supplier.tags) if supplier.tags else ""

            data = [
                supplier.id,
                supplier.company_name,
                supplier.url,
                supplier.company_type or "",
                supplier.city or "",
                supplier.country or "",
                supplier.email or "",
                supplier.phone or "",
                tags_str,
                supplier.created_at.strftime('%Y-%m-%d %H:%M:%S') if supplier.created_at else "",
                supplier.updated_at.strftime('%Y-%m-%d %H:%M:%S') if supplier.updated_at else ""
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

            # 设置行高（自动换行）
            ws.row_dimensions[row_num].height = 20

        # 冻结首行
        ws.freeze_panes = "A2"

        # 添加筛选器
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        logger.info(f"Excel文件生成完成，共 {total} 行数据")

        # 保存到内存
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def get_filename(prefix: str = "suppliers") -> str:
        """
        生成Excel文件名

        Args:
            prefix: 文件名前缀

        Returns:
            str: 文件名
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{prefix}_{timestamp}.xlsx"
