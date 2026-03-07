"""供应商数据导出到Excel"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.database import SessionLocal
from api.models import Supplier

EXPORT_DIR = Path("/root/.openclaw/workspace/exports/excel")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def create_styled_cell(worksheet, cell, font_bold=False, fill_color=None):
    """创建样式化的单元格"""
    font = Font(name='微软雅黑', size=11, bold=font_bold)
    cell.font = font

    if fill_color:
        cell.fill = PatternFill(start_type='solid', fgColor=fill_color)
    else:
        cell.fill = PatternFill(start_type='solid', fgColor='FFFFFF')

    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border


def export_to_excel():
    """导出供应商数据到Excel"""
    print("📊 查询数据库...")
    db = SessionLocal()

    try:
        suppliers = db.query(Supplier).order_by(Supplier.created_at.desc()).all()
        print(f"✅ 找到 {len(suppliers)} 家供应商")

        wb = Workbook()
        ws = wb.active
        ws.title = "供应商数据"

        headers = [
            "ID", "公司标题", "网址", "域名",
            "邮箱", "电话", "地址",
            "城市", "国家", "公司类型",
            "中文描述", "标签",
            "来源", "发现日期", "创建时间"
        ]

        column_widths = {
            'A': 6, 'B': 30, 'C': 30, 'D': 20,
            'E': 30, 'F': 20, 'G': 30, 'H': 10,
            'I': 8, 'J': 12, 'K': 40, 'L': 20,
            'M': 10, 'N': 15, 'O': 20,
        }

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            create_styled_cell(ws, cell, font_bold=True, fill_color='E0E0E0')

        for row_num, supplier in enumerate(suppliers, 2):
            emails_str = ', '.join(supplier.emails) if supplier.emails else ''
            phones_str = ', '.join(supplier.phones) if supplier.phones else ''
            tags_str = ', '.join(supplier.tags) if supplier.tags else ''

            date_found_str = supplier.date_found.strftime('%Y-%m-%d') if supplier.date_found else ''
            created_at_str = supplier.created_at.strftime('%Y-%m-%d %H:%M') if supplier.created_at else ''

            ws.cell(row=row_num, column=1).value = supplier.id
            ws.cell(row=row_num, column=2).value = supplier.title
            ws.cell(row=row_num, column=3).value = supplier.url
            ws.cell(row=row_num, column=4).value = supplier.domain
            ws.cell(row=row_num, column=5).value = emails_str
            ws.cell(row=row_num, column=6).value = phones_str
            ws.cell(row=row_num, column=7).value = supplier.address
            ws.cell(row=row_num, column=8).value = supplier.city
            ws.cell(row=row_num, column=9).value = supplier.country
            ws.cell(row=row_num, column=10).value = supplier.company_type
            ws.cell(row=row_num, column=11).value = supplier.description_cn
            ws.cell(row=row_num, column=12).value = tags_str
            ws.cell(row=row_num, column=13).value = supplier.source
            ws.cell(row=row_num, column=14).value = date_found_str
            ws.cell(row=row_num, column=15).value = created_at_str

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"suppliers_{timestamp}.xlsx"
        filepath = EXPORT_DIR / filename

        wb.save(str(filepath))
        print(f"✅ 导出成功: {filepath}")
        print(f"📦 文件大小: {filepath.stat().st_size / 1024:.2f} KB")
        return str(filepath)

    except Exception as e:
        print(f"❌ 导出失败: {e}")
        return False
    finally:
        db.close()


if __name__ == "__main__":
    export_to_excel()
